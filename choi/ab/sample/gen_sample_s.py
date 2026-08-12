# -*- coding: utf-8 -*-
"""サンプルS生成: AB会話教材(ID1-5) → ミニブック投稿形式 14列xlsx + audio.zip + image.zip
根拠: 設計12 サンプルデータ作成指示書 / 投稿データ_メディアパック仕様.md
"""
import json, re, sys, os, zipfile, subprocess
import openpyxl
from openpyxl import Workbook

SRC = "/home/claude/source_ab.xlsx"
MANIFEST = "/mnt/user-data/uploads/manifest.json"
OUT = "/home/claude/out"
os.makedirs(OUT, exist_ok=True)

errors = []
def check(cond, msg):
    if not cond:
        errors.append(msg)
        print("NG:", msg)

# ---------- 1. 元データ読込 (ID 1-5) ----------
wb = openpyxl.load_workbook(SRC)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
assert header == ('ID','カテゴリー','テーマ','2会話','類似フレーズA','類似フレーズB','AB会話の話者プロファイル'), header
src = [r for r in rows[1:] if r[0] is not None and int(r[0]) in (1,2,3,4,5)]
check(len(src) == 5, f"ID1-5 の行数 {len(src)} != 5")

# ---------- 2. パース関数 ----------
def parse_joyaku(line):
    """（…）を剥がし U+3000→半角スペース1個、圧縮・トリム"""
    m = re.fullmatch(r'（(.*)）', line.strip())
    assert m, f"語順訳の括弧不正: {line!r}"
    v = m.group(1)
    v = re.sub(r'[\u3000 ]+', ' ', v).strip()
    return v

def parse_2kaiwa(cell):
    lines = [l for l in cell.split('\n') if l.strip()]
    assert len(lines) == 4, f"2会話は4行想定: {lines!r}"
    assert lines[0].startswith('A：') and lines[2].startswith('B：'), lines
    a_en = lines[0][2:].strip()
    a_jo = parse_joyaku(lines[1])
    b_en = lines[2][2:].strip()
    b_jo = parse_joyaku(lines[3])
    return (a_en, a_jo), (b_en, b_jo)

def parse_ruiji(cell):
    lines = [l for l in cell.split('\n') if l.strip()]
    assert len(lines) == 4, f"類似は4行想定: {lines!r}"
    return (lines[0].strip(), parse_joyaku(lines[1])), (lines[2].strip(), parse_joyaku(lines[3]))

def parse_profile(cell):
    lines = [l for l in cell.split('\n') if l.strip()]
    assert len(lines) == 2, f"プロファイルは2行想定: {lines!r}"
    assert lines[0].startswith('A: ') and lines[1].startswith('B: '), lines
    return lines[0][3:].strip(), lines[1][3:].strip()

def parse_theme(cell):
    lines = [l for l in cell.split('\n') if l.strip()]
    assert len(lines) == 2, f"テーマは2行想定: {lines!r}"
    return lines[0].strip(), lines[1].strip()  # en, ja

# ---------- 3. 行生成 ----------
themes = set()
category = None
IMG = "theme001.png"
data_rows = []   # dicts
audio_names = []
row_meta = []    # (set_no, role, num, en) 検査9用
for r in src:
    sid, cat, theme, kaiwa, ruiA, ruiB, prof = r
    sid = int(sid)
    category = cat
    t_en, t_ja = parse_theme(theme)
    themes.add((t_en, t_ja))
    (a0, a0j), (b0, b0j) = parse_2kaiwa(kaiwa)
    (a1, a1j), (a2, a2j) = parse_ruiji(ruiA)
    (b1, b1j), (b2, b2j) = parse_ruiji(ruiB)
    pA, pB = parse_profile(prof)
    card_attr = json.dumps({"speakers": {"A": pA, "B": pB}},
                           ensure_ascii=False, separators=(',', ':'))
    six = [
        ('A',    a0, a0j, 'a', 0),
        ('B',    b0, b0j, 'b', 0),
        ('類似A', a1, a1j, 'a', 1),
        ('類似B', b1, b1j, 'b', 1),
        ('類似A', a2, a2j, 'a', 2),
        ('類似B', b2, b2j, 'b', 2),
    ]
    for i, (spk, en, jo, role, num) in enumerate(six):
        fname = f"id{sid:03d}_{role}{num}.mp3"
        audio_names.append(fname)
        row_meta.append((sid, role, num, en))
        data_rows.append({
            'チャプター名': t_ja,
            'シーン名': f"セット{sid}",
            '話者': spk,
            '英会話': en,
            '語順訳': jo,
            '和訳': '',
            '音声ファイル': fname,
            'デッキ画像': IMG if (sid == 1 and i == 0) else '',
            'カード画像': IMG if i == 0 else '',
            'カード属性': card_attr if i == 0 else '',
            'デッキ属性': '' , # 後で先頭行に設定
        })

check(len(themes) == 1, f"ID1-5 のテーマが単一でない: {themes}")
t_en, t_ja = next(iter(themes))
deck_attr = json.dumps({"title_en": t_en, "title_ja": t_ja},
                       ensure_ascii=False, separators=(',', ':'))
data_rows[0]['デッキ属性'] = deck_attr
print(f"カテゴリー(ユニット名) = {category}")
print(f"デッキ = {t_ja} / {t_en}, 行数 = {len(data_rows)}")

# ---------- 4. xlsx 出力 (14列) ----------
# 注: 指示書§2は「語順訳」だが、根拠仕様§1.2の既知ヘッダーは
# 「手順3で作成した語順訳」。§4.1により未知ヘッダーは無警告で読み飛ばされる
# ため、正本ヘッダーを採用(本文で補正報告)。
JOYAKU_HEADER = "手順3で作成した語順訳"
HEADERS = ["デッキ編集ID","カード編集ID","センテンス編集ID","チャプター名","シーン名",
           "話者","英会話",JOYAKU_HEADER,"和訳","音声ファイル","デッキ画像","カード画像",
           "カード属性","デッキ属性"]
out_wb = Workbook()
ows = out_wb.active
ows.title = "Sheet1"
for c, h in enumerate(HEADERS, 1):
    cell = ows.cell(row=1, column=c, value=h)
    cell.number_format = '@'
for ri, d in enumerate(data_rows, 2):
    vals = ['', '', '', d['チャプター名'], d['シーン名'], d['話者'], d['英会話'],
            d['語順訳'], d['和訳'], d['音声ファイル'], d['デッキ画像'], d['カード画像'],
            d['カード属性'], d['デッキ属性']]
    for c, v in enumerate(vals, 1):
        if v == '':
            continue  # 空欄セルは値なし
        cell = ows.cell(row=ri, column=c, value=str(v))
        cell.number_format = '@'
xlsx_path = os.path.join(OUT, "sampleS_unit.xlsx")
out_wb.save(xlsx_path)
print("wrote", xlsx_path)

# ---------- 5. 出口検査 ----------
wb2 = openpyxl.load_workbook(xlsx_path)
w2 = wb2["Sheet1"]
rows2 = list(w2.iter_rows(values_only=False))
hdr2 = [c.value for c in rows2[0]]
data2 = rows2[1:]
col = {h: i for i, h in enumerate(hdr2)}

# 検査1: 行数と話者パターン
check(len(data2) == 5 * 6, f"行数 {len(data2)} != 30")
expected_spk = ['A', 'B', '類似A', '類似B', '類似A', '類似B']
for cidx in range(5):
    block = data2[cidx*6:(cidx+1)*6]
    spks = [r[col['話者']].value for r in block]
    check(spks == expected_spk, f"カード{cidx+1} 話者順 {spks}")

# 検査2: スラッシュ数 = 半角スペース数
for i, r in enumerate(data2):
    en = r[col['英会話']].value or ''
    jo = r[col[JOYAKU_HEADER]].value or ''
    check(en.count('/') == jo.count(' '),
          f"行{i+2}: slash={en.count('/')} space={jo.count(' ')} : {en!r} / {jo!r}")

# 検査3: 正規化冪等性(U+3000/改行/連続空白/前後空白なし)
bad_ws = re.compile(r'\u3000|\n|\r|  |\t')
for ri, r in enumerate(data2, 2):
    for c in r:
        if c.value is None:
            continue
        v = str(c.value)
        check(not bad_ws.search(v), f"行{ri}列{c.column} 空白/改行不正: {v!r}")
        check(v == v.strip(), f"行{ri}列{c.column} 前後空白: {v!r}")

# 検査4: 全セル文字列型
for ri, r in enumerate(data2, 2):
    for c in r:
        if c.value is not None:
            check(isinstance(c.value, str), f"行{ri}列{c.column} 非文字列型: {type(c.value)}")

# 検査5: 拡張属性 先頭行のみ / JSONパース可 / ミニファイ一致
for name, allowed in (('カード属性', {2, 8, 14, 20, 26}), ('デッキ属性', {2})):
    for ri, r in enumerate(data2, 2):
        v = r[col[name]].value
        if ri in allowed:
            check(v is not None, f"{name} 行{ri} 先頭行が空欄")
            try:
                parsed = json.loads(v)
                mini = json.dumps(parsed, ensure_ascii=False, separators=(',', ':'))
                check(mini == v, f"{name} 行{ri} ミニファイ不一致")
            except Exception as e:
                check(False, f"{name} 行{ri} JSONパース不能: {e}")
        else:
            check(v is None, f"{name} 行{ri} 先頭行以外に記入: {v!r}")

# 検査6: シーン名の境界と連番
scenes = [r[col['シーン名']].value for r in data2]
check(scenes == [f"セット{i}" for i in range(1, 6) for _ in range(6)], f"シーン名列不正: {scenes}")

# 検査7: 音声ファイル列
audio_cells = [r[col['音声ファイル']].value for r in data2]
check(all(audio_cells), "音声ファイル列に空欄あり")
check(len(set(audio_cells)) == len(audio_cells) == 30, "音声参照の件数/重複不正")
pat = re.compile(r'^[A-Za-z0-9_\-\.]+$')
for a in audio_cells:
    check(bool(pat.fullmatch(a)), f"音声ファイル名文字集合違反: {a}")

# 検査8: 画像列
for ri, r in enumerate(data2, 2):
    dv, cv = r[col['デッキ画像']].value, r[col['カード画像']].value
    check((dv == IMG) if ri == 2 else dv is None, f"デッキ画像 行{ri}: {dv!r}")
    check((cv == IMG) if ri in {2, 8, 14, 20, 26} else cv is None, f"カード画像 行{ri}: {cv!r}")

# 検査9: 英文正規化 = manifest キー(30件全件)
manifest = json.load(open(MANIFEST, encoding='utf-8'))['items']
def norm(en):
    v = en.replace('/', ' ')
    v = re.sub(r'\s+', ' ', v).strip().lower()
    return v
rename_map = []  # (old, new)
for (sid, role, num, en), fname in zip(row_meta, audio_names):
    key = norm(en)
    check(key in manifest, f"manifestにキーなし: {key!r}")
    if key in manifest:
        old = manifest[key]['file']
        expect_old = f"s{sid}_{role}{num}.mp3"
        check(old == expect_old, f"manifest対応不一致: {key!r} -> {old} (期待 {expect_old})")
        rename_map.append((old, fname, key))
check(len(rename_map) == 30, "リネーム対応が30件でない")

# 検査10: ハングル
hangul = re.compile(r'[\uAC00-\uD7AF\u1100-\u11FF\u3130-\u318F\uA960-\uA97F\uD7B0-\uD7FF]')
for ri, r in enumerate(rows2, 1):
    for c in r:
        if c.value and hangul.search(str(c.value)):
            check(False, f"ハングル混入 行{ri}列{c.column}")

# 編集ID列: ヘッダーのみ・全行空欄
for ri, r in enumerate(data2, 2):
    for h in ("デッキ編集ID", "カード編集ID", "センテンス編集ID"):
        check(r[col[h]].value is None, f"{h} 行{ri} に値あり")

# ---------- 6. audio.zip (プレースホルダ無音) ----------
sil = os.path.join(OUT, "_silence.mp3")
subprocess.run(["ffmpeg", "-y", "-f", "lavfi", "-i", "anullsrc=r=24000:cl=mono",
                "-t", "0.5", "-b:a", "32k", "-loglevel", "error", sil], check=True)
azip = os.path.join(OUT, "audio.zip")
with zipfile.ZipFile(azip, 'w', zipfile.ZIP_STORED) as z:
    for a in sorted(set(audio_cells)):
        z.write(sil, a)
with zipfile.ZipFile(azip) as z:
    names = z.namelist()
    check(sorted(names) == sorted(audio_cells), "audio.zip 集合不一致")
    check(all('/' not in n for n in names), "audio.zip 非フラット")
print("wrote", azip, len(names), "files (placeholder silence)")

# リネーム対応表・実音声zip作成スクリプト
with open(os.path.join(OUT, "audio_rename_map.tsv"), 'w', encoding='utf-8') as f:
    f.write("old_name\tnew_name\tmanifest_key\n")
    for old, new, key in rename_map:
        f.write(f"{old}\t{new}\t{key}\n")
with open(os.path.join(OUT, "make_real_audio_zip.sh"), 'w', encoding='utf-8') as f:
    f.write("#!/bin/sh\n# 使い方: sh make_real_audio_zip.sh <実mp3のあるディレクトリ>\n")
    f.write("# s*_*.mp3 を id*_*.mp3 にコピーして audio.zip を作る\n")
    f.write('SRC="$1"; T=$(mktemp -d) || exit 1\n')
    for old, new, _ in rename_map:
        f.write(f'cp "$SRC/{old}" "$T/{new}" || exit 1\n')
    f.write('(cd "$T" && zip -X -q audio.zip id*.mp3) && mv "$T/audio.zip" ./audio.zip\n')
    f.write('rm -rf "$T"; echo "audio.zip written ($(unzip -l audio.zip | tail -1))"\n')

# ---------- 7. image.zip ----------
from PIL import Image, ImageDraw, ImageFont
W, H = 960, 540
img = Image.new('RGB', (W, H))
top, bottom = (79, 70, 229), (37, 99, 235)  # indigo->blue (プロトタイプ配色系)
for y in range(H):
    t = y / H
    img.paste(tuple(int(a + (b - a) * t) for a, b in zip(top, bottom)), (0, y, W, y + 1))
d = ImageDraw.Draw(img)
font = None
for fp in ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",):
    if os.path.exists(fp):
        font = ImageFont.truetype(fp, 54)
        small = ImageFont.truetype(fp, 28)
if font:
    d.text((W//2, H//2 - 30), "English for Yoroshiku", anchor="mm", font=font, fill=(255,255,255))
    d.text((W//2, H//2 + 40), "AB Conversation  /  Theme 001", anchor="mm", font=small, fill=(219,234,254))
img_path = os.path.join(OUT, IMG)
img.save(img_path)
izip = os.path.join(OUT, "image.zip")
with zipfile.ZipFile(izip, 'w', zipfile.ZIP_DEFLATED) as z:
    z.write(img_path, IMG)
print("wrote", izip)

# ---------- 結果 ----------
print("\n=== 出口検査結果 ===")
if errors:
    print(f"NG {len(errors)} 件")
    sys.exit(1)
print("全10項目 + 追加検査 パス (エラー0件)")
